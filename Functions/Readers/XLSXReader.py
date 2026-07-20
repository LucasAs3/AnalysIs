from openpyxl import load_workbook

def excel_reader(file: str) -> str:
    doc = load_workbook(file)
    plan = doc.active


    linhas = []

    for linha in plan.iter_rows(values_only=True):
        line = " | ".join(str(x) for x in linha if x is not None)
        linhas.append(line)

    doc.close()

    return "\n".join(linhas)